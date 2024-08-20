import re
import pdfplumber
import pandas as pd
from collections import namedtuple
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font

# Define namedtuples for different scenarios
Line = namedtuple('Line', 'Commercial_Package_Policy Premium_Policy Premium_Policy1')
CoverageDetails = namedtuple('CoverageDetails', 'Coverage Premium1 Premium2 Premium3 Premium1_1 Premium2_1 Premium3_1')

# List of PDF files to process
files1 = ['7.pdf', '7.pdf']  # PDF files for the first part
# files2 = ['6.pdf', '6.pdf']  # PDF files for the second part

# Regex patterns for part 1
amount_re = re.compile(r'([A-Za-z\s,]+?)(?:\s+\$([\d,]+\.\d{2})| Not Covered|:\s*\$.00)')
total_re = re.compile(r'Estimated Total Premium:\s*\$([\d,]+\.\d{2})')
amount_re2 = re.compile(r'([A-Za-z\s$$]+)\s+\$([\d,]+(?:\.\d{2})?)')
amount_re1 = re.compile(r'([A-Za-z\s/()]+)\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?\s+(\$\d{1,3}(?:,\d{3})*(?:/\s*\$?\d{1,3}(?:,\d{3})*)?)?(?:\s+Included)?')
total_re1 = re.compile(r'([A-Za-z\s,]+?)(?:\s+\$([\d,]+\.\d{2})(?:\s+\$([\d,]+\.\d{2}))(?:\s+\$([\d,]+\.\d{2})))')

# Initialize lists to hold the extracted data for part 1
lines_list1 = []
lines_list_first_pdf = []  
lines_list_second_pdf = []
# Process the first PDF file for part 1
with pdfplumber.open(files1[0]) as pdf:  
    for page in pdf.pages:  
        text = page.extract_text()  
        if text:  
            lines = text.split('\n')
            for line in lines:
                match = amount_re1.search(line)
                match1 = total_re1.search(line)
                if match:
                    coverage = match.group(1).strip()  
                    premium1 = match.group(2).strip() if match.group(2) else None  
                    premium2 = match.group(3).strip() if match.group(3) else None  
                    premium3 = match.group(4).strip() if match.group(4) else None  
                    lines_list_first_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
                elif match1:
                    coverage = match1.group(1).strip()
                    premium1 = match1.group(2).strip() if match1.group(2) else None
                    premium2 = match1.group(3).strip() if match1.group(3) else None
                    premium3 = match1.group(4).strip() if match1.group(4) else None
                    lines_list_first_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
            #print("Extracted text for PDF 1:", lines)  # Debug print
            
            # Processing coverage parts
            index = [i for i, x in enumerate(lines) if "This policy consists of the following coverage parts: " in x]
            if index:
                start_line = index[0] + 1
                for line in lines[start_line:]:
                    match2 = amount_re.search(line)
                    if match2:
                        coverage = match2.group(1).strip()
                        premium = match2.group(2).strip() if match2.group(2) else "Not Covered"
                        lines_list1.append(Line(Commercial_Package_Policy=coverage, Premium_Policy=premium, Premium_Policy1=""))
                        #print(f"Matched coverage: {coverage}, premium: {premium}")  # Debug print

                    match3 = total_re.search(line)
                    if match3:
                        total_premium = match3.group(1).strip()
                        lines_list1.append(Line("Estimated Total Premium", total_premium, ""))
                        #print(f"Matched total premium: {total_premium}")  # Debug print

            index1 = [i for i, x in enumerate(lines) if "Primary use of the vehicle: Pleasure/Personal" in x]
            if index1:
                start_line1 = index1[0] + 1
                for line in lines[start_line1:]:
                    match = amount_re2.search(line)
                    if match:
                        coverage = match.group(1).strip()
                        premium = match.group(2).strip() if match.group(2) else "Not Covered"
                        lines_list1.append(Line(Commercial_Package_Policy=coverage, Premium_Policy=premium, Premium_Policy1=""))
                        #print(f"Matched additional coverage: {coverage}, premium: {premium}")  # Debug print

# Process the second PDF file and update Premium1 for part 1
with pdfplumber.open(files1[1]) as pdf:  
    for page in pdf.pages:  
        text = page.extract_text()  
        if text:  
            lines = text.split('\n')
            for line in lines:
                match = amount_re1.search(line)
                match1 = total_re1.search(line)
                if match:
                    coverage = match.group(1).strip()  
                    premium1 = match.group(2).strip() if match.group(2) else None  
                    premium2 = match.group(3).strip() if match.group(3) else None  
                    premium3 = match.group(4).strip() if match.group(4) else None  
                    lines_list_second_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
                elif match1:
                    coverage = match1.group(1).strip()
                    premium1 = match1.group(2).strip() if match1.group(2) else None
                    premium2 = match1.group(3).strip() if match1.group(3) else None
                    premium3 = match1.group(4).strip() if match1.group(4) else None
                    lines_list_second_pdf.append(CoverageDetails(Coverage=coverage, Premium1=premium1, Premium2=premium2, Premium3=premium3, Premium1_1=None, Premium2_1=None, Premium3_1=None))
            #print("Extracted text for PDF 2:", lines)  # Debug print

            # Processing coverage parts
            index = [i for i, x in enumerate(lines) if "This policy consists of the following coverage parts: " in x]
            if index:
                start_line = index[0] + 1
                for i, line in enumerate(lines[start_line:]):
                    match2 = amount_re.search(line)
                    if match2 and i < len(lines_list1):  
                        premium1 = match2.group(2).strip() if match2.group(2) else "Not Covered"
                        lines_list1[i] = lines_list1[i]._replace(Premium_Policy1=premium1)  
                        #print(f"Updated line {i} with Premium_Policy1: {premium1}")  # Debug print

                    match3 = total_re.search(line)
                    if match3 and i < len(lines_list1):  
                        total_premium1 = match3.group(1).strip()
                        if lines_list1[i].Commercial_Package_Policy == "Estimated Total Premium":
                            lines_list1[i] = lines_list1[i]._replace(Premium_Policy1=total_premium1)  
                            #print(f"Updated total premium for Estimated Total Premium: {total_premium1}")  # Debug print

            index1 = [i for i, x in enumerate(lines) if "Primary use of the vehicle: Pleasure/Personal" in x]
            if index1:
                start_line1 = index1[0] + 1
                for line in lines[start_line1:]:
                    match = amount_re2.search(line)
                    if match:
                        coverage = match.group(1).strip()
                        premium1 = match.group(2).strip() if match.group(2) else "Not Covered"
                        matched = False
                        for i, entry in enumerate(lines_list1):
                            if entry.Commercial_Package_Policy == coverage and not entry.Premium_Policy1:
                                lines_list1[i] = lines_list1[i]._replace(Premium_Policy1=premium1)
                                matched = True
                                break
                        if not matched:
                            lines_list1.append(Line(Commercial_Package_Policy=coverage, Premium_Policy="", Premium_Policy1=premium1))
                            #print(f"Added new coverage: {coverage}, Premium_Policy1: {premium1}")  # Debug print

# Convert to DataFrame for part 1
coverage_df1 = pd.DataFrame(lines_list1)

# Verify DataFrame columns and data
# print("DataFrame for part 1:", coverage_df1)

# Check if DataFrame is empty
if coverage_df1.empty:
    print("No data was collected for coverage_df1.")
else:
    # Handle empty strings and non-numeric values in 'Premium' and 'Premium1' for part 1
    coverage_df1['Premium_Policy'] = coverage_df1['Premium_Policy'].replace(['Not Covered', ''], '0').str.replace(',', '').astype(float)
    coverage_df1['Premium_Policy1'] = coverage_df1['Premium_Policy1'].replace(['Not Covered', ''], '0').str.replace(',', '').astype(float)

    # Add a new column 'Difference' by subtracting Premium1 from Premium for part 1
    coverage_df1['Difference'] = coverage_df1['Premium_Policy'] - coverage_df1['Premium_Policy1']
coverage_df2 = pd.DataFrame(lines_list_first_pdf)

# Verify DataFrame columns for coverage_df2
# print("DataFrame columns for part 2:", coverage_df2.columns)

# Match and update the DataFrame with data from the second PDF for part 2
for entry2 in lines_list_second_pdf:
    matched = False
    for i, row in coverage_df2.iterrows():
        if row['Coverage'] == entry2.Coverage:
            coverage_df2.at[i, 'Premium1_1'] = entry2.Premium1
            coverage_df2.at[i, 'Premium2_1'] = entry2.Premium2
            coverage_df2.at[i, 'Premium3_1'] = entry2.Premium3
            matched = True
            break
    if not matched:
        print(f"Coverage '{entry2.Coverage}' from second PDF did not match any entry in the first PDF.")

# Add columns for differences for part 2
coverage_df2['difference1_1'] = None
coverage_df2['difference2_1'] = None
coverage_df2['difference3_1'] = None

# Calculate differences for part 2
for i, row in coverage_df2.iterrows():
    def parse_premium(premium_str):
        if premium_str is None:
            return None
        numeric_parts = re.findall(r'\d+(?:,\d{3})*(?:\.\d{2})?', premium_str)
        if numeric_parts:
            return float(numeric_parts[0].replace(',', ''))
        return None

    premium1 = parse_premium(row['Premium1'])
    premium1_1 = parse_premium(row['Premium1_1'])
    premium2 = parse_premium(row['Premium2'])
    premium2_1 = parse_premium(row['Premium2_1'])
    premium3 = parse_premium(row['Premium3'])
    premium3_1 = parse_premium(row['Premium3_1'])
    
    coverage_df2.at[i, 'difference1_1'] = (premium1_1 - premium1) if premium1 is not None and premium1_1 is not None else None
    coverage_df2.at[i, 'difference2_1'] = (premium2_1 - premium2) if premium2 is not None and premium2_1 is not None else None
    coverage_df2.at[i, 'difference3_1'] = (premium3_1 - premium3) if premium3 is not None and premium3_1 is not None else None

# Display the updated DataFrames
print("Coverage DataFrame from part 1:")
print(coverage_df1)
print("\nCoverage DataFrame from part 2:")
print(coverage_df2)
#
# ... [Previous code remains unchanged] ...

excel_file = 'comparison.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    # Write the policy data to Excel
    coverage_df1.to_excel(writer, index=False, sheet_name='Comparison Report', startrow=2)
    
    # Get the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Comparison Report']

    # Add headings for the policy report
    report_heading = "Comparison Report"
    worksheet.merge_cells('A1:D1')  # Merge cells for the heading
    cell = worksheet.cell(row=1, column=1)
    cell.value = report_heading
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    
    # Set minimum column width
    min_width = 35
    
    # Set column widths and apply border styles for the policy report
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in col) + 2  # Adding 2 for padding
        col_letter = get_column_letter(col[0].column)  # Get column letter
        adjusted_width = max(min_width, max_length)  # Ensure minimum width
        worksheet.column_dimensions[col_letter].width = adjusted_width  # Set column width
        for cell in col:
            cell.border = border

    # Define start_row after the first DataFrame
    start_row = len(coverage_df1) + 4  # Start 2 rows below the last premium row
    
    # Write the second DataFrame to Excel starting at the calculated start_row
    coverage_df2.to_excel(writer, index=False, sheet_name='Comparison Report', startrow=start_row, header=True)

    # Add headings for the premises report
    premises_heading = "Comparison Report"
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)  # Merge cells for the heading
    cell = worksheet.cell(row=start_row, column=1)
    cell.value = premises_heading
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)

    # Set column widths and apply border styles for the premises report
    for col in worksheet.iter_cols(min_row=start_row, max_row=start_row+len(coverage_df2)+1):
        max_length = max(len(str(cell.value)) for cell in col) + 2  # Adding 2 for padding
        col_letter = get_column_letter(col[0].column)  # Get column letter
        adjusted_width = max(min_width, max_length)  # Ensure minimum width
        worksheet.column_dimensions[col_letter].width = adjusted_width  # Set column width
        for cell in col:
            cell.border = border

# Save the workbook
workbook.save(excel_file)
print(f'Data successfully written to {excel_file}')